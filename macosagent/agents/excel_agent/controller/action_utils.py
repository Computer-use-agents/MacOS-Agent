
import re
def escape_single_quotes(text):
    # 匹配未转义的单引号（不匹配 \\'）
    pattern = r"(?<!\\)'"
    return re.sub(pattern, r"\\'", text)



def parsing_response_to_pyautogui_code(responses, input_swap:bool=True) -> str:
    '''
    将M模型的输出解析为OSWorld中的action，生成pyautogui代码字符串
    参数:
        response: 包含模型输出的字典，结构类似于：
        {
            "action_type": "hotkey",
            "action_inputs": {
                "hotkey": "command v",
                "start_box": None,
                "end_box": None
            }
        }
    返回:
        生成的pyautogui代码字符串
    '''

    pyautogui_code = f"import pyautogui\nimport time\n"
    if isinstance(responses, dict):
        responses = [responses]
    for response_id, response in enumerate(responses):
        if "observation" in response:
            observation = response["observation"]
        else:
            observation = ""

        if "thought" in response:
            thought = response["thought"]
        else:
            thought = ""
        
        # if response_id == 0:
        #     pyautogui_code += f"'''\nObservation:\n{observation}\n\nThought:\n{thought}\n'''\n"
        # else:
        #     pyautogui_code += f"\ntime.sleep(3)\n"

        action_dict = response
        action_type = action_dict.get("action_type")
        action_inputs = action_dict.get("action_inputs", {})
        
        if action_type == "HOT_KEY":
            # Parsing hotkey action
            hotkey = action_inputs.get("value", "").strip()
            if hotkey:
                # Handle other hotkeys
                keys = hotkey.split(" ")  # Split the keys by space
                pyautogui_code += f"\ntime.sleep(0.5)\n"
                pyautogui_code += f"\npyautogui.hotkey({', '.join([repr(k) for k in keys])})"
            else:
                pyautogui_code += f"\n# Unrecognized hot key action parsing: {action_type}"
                logger.warning(f"Unrecognized hot key action parsing: {response}")
        elif action_type == "ENTER":
            pyautogui_code += f"\npyautogui.press('enter')"
        elif action_type == "INPUT":
            # Parsing typing action using clipboard
            content = action_inputs.get("value", "")
            content = escape_single_quotes(content)
            if content:
                if input_swap:
                    pyautogui_code += f"\nimport pyperclip"
                    pyautogui_code += f"\npyperclip.copy('{content.strip()}')"
                    pyautogui_code += f"\ntime.sleep(0.5)\n"
                    pyautogui_code += f"\npyautogui.hotkey('command', 'v')"
                    pyautogui_code += f"\ntime.sleep(0.5)\n"
                    if content.endswith("\n") or content.endswith("\\n"):
                        pyautogui_code += f"\npyautogui.press('enter')"
                else:
                    pyautogui_code += f"\npyautogui.write('{content.strip()}', interval=0.1)"
                    pyautogui_code += f"\ntime.sleep(0.5)\n"
                    if content.endswith("\n") or content.endswith("\\n"):
                        pyautogui_code += f"\npyautogui.press('enter')"

        
        elif action_type in ["DRAG", "SELECT"]:
            # Parsing drag or select action based on start and end_boxes
            points = action_inputs.get("position")
            assert len(points) == 2 and len(points[0]) == 2 and len(points[1]) == 2, "Drag or select action should have 2 points"
            start_box = points[0]
            end_box = points[1]
            
            if start_box and end_box:
                x1, y1 = start_box # [x, y] float number between 0 and 1
                # sx = round(float(x1) * image_width, 3)
                # sy = round(float(y1) * image_height, 3)
                sx = round(float(x1), 3)
                sy = round(float(y1), 3)
                x2, y2 = end_box # [x, y] float number between 0 and 1
                # ex = round(float(x2) * image_width, 3)
                # ey = round(float(y2) * image_height, 3)
                ex = round(float(x2), 3)
                ey = round(float(y2), 3)
                pyautogui_code += (
                    f"\npyautogui.moveTo({sx}, {sy})\n"
                    f"\npyautogui.dragTo({ex}, {ey}, duration=1.0)\n"
                )
            else:
                logger.warning(f"Parse failed! Drag or select action should have 2 points: {response}")
        elif action_type == "SCROLL":
            # Parsing scroll action
            start_box = action_inputs.get("position")
            if start_box:
                # x1, y1, x2, y2 = eval(start_box)  # Assuming box is in [x1, y1, x2, y2]
                # x = round(float((x1 + x2) / 2) * image_width, 3)
                # y = round(float((y1 + y2) / 2) * image_height, 3)
                x, y = start_box
                # x = round(float(x) * image_width, 3)
                # y = round(float(y) * image_height, 3)
                x = round(float(x), 3)
                y = round(float(y), 3)                
                # # 先点对应区域，再滚动
                pyautogui_code += f"\npyautogui.click({x}, {y}, button='left')"
            else:
                x = None
                y = None
            direction = action_inputs.get("direction", "")
            
            if x == None:
                if "up" in direction.lower():
                    pyautogui_code += f"\npyautogui.scroll(5)"
                elif "down" in direction.lower():
                    pyautogui_code += f"\npyautogui.scroll(-5)"
            else:
                if "up" in direction.lower():
                    pyautogui_code += f"\npyautogui.scroll(5, x={x}, y={y})"
                elif "down" in direction.lower():
                    pyautogui_code += f"\npyautogui.scroll(-5, x={x}, y={y})"

        elif action_type in ["CLICK", "LEFT_CLICK_DOUBLE", "LEFT_CLICK_SINGLE", "RIGHT_CLICK_SINGLE", "HOVER"]:
            # Parsing mouse click actions
            start_box = action_inputs.get("position")
            # start_box = str(start_box)
            if start_box:
                # start_box = eval(start_box)
                if len(start_box) == 2:
                    x1, y1 = start_box
                    x2 = x1
                    y2 = y1
                else:
                    raise ValueError(f"Unknown start_box format: {start_box}")
                # x = round(float((x1 + x2) / 2) * image_width, 3)
                # y = round(float((y1 + y2) / 2) * image_height, 3)

                x = round(float((x1 + x2)/2), 3)
                y = round(float((y1 + y2)/2), 3)

                if action_type == "LEFT_CLICK_DOUBLE" or action_type == "CLICK":
                    pyautogui_code += f"\npyautogui.click({x}, {y}, button='left')"
                elif action_type == "LEFT_CLICK_DOUBLE":
                    pyautogui_code += f"\npyautogui.doubleClick({x}, {y}, button='left')"
                elif action_type == "RIGHT_CLICK_SINGLE":
                    pyautogui_code += f"\npyautogui.click({x}, {y}, button='right')"
                elif action_type == "HOVER":
                    pyautogui_code += f"\npyautogui.moveTo({x}, {y})"
        
        elif action_type in ["FINISH"]:
            pyautogui_code = f"DONE"
        
        else:
            pyautogui_code += f"\n# Unrecognized action type: {action_type}"
            logger.error(f"Unrecognized action type: {response}")

    return pyautogui_code

import openpyxl
from openpyxl.utils import get_column_letter


def insert_text_to_cell_function(file_path, file_save_path, sheet_name, cell_address,insert_text, insert_pos):
    """
    在Excel单元格的指定字符位置后插入文本
    
    参数说明：
    - file_path:   Excel文件路径（不存在则自动创建）
    - sheet_name:  工作表名称（不存在则自动创建）
    - cell_address: 目标单元格地址（如"A1"）
    - insert_pos:   插入位置（从0开始，0=开头）
    - insert_text:  待插入的文本
    
    返回值：
    - 字典包含三个键：
      - success:      操作是否成功（True/False）
      - error_message: 错误描述（成功时为空）
      - result_text:   插入后的完整文本（失败时为空）
    """
    result = {"success": False, "error_message": "", "result_text": ""}
    
    try:
        # 加载或创建Excel文件
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
        
        # 获取单元格并处理空值
        cell = sheet[cell_address]
        original_text = str(cell.value) if cell.value is not None else ""
        
        # 处理插入位置为负数的情况
        if insert_pos < 0:
            result["error_message"] = f"错误：插入位置不能为负数（当前值：{insert_pos}）"
            return result
        
        # 空单元格处理逻辑
        if original_text == "":
            if insert_pos != 0:
                result["error_message"] = "错误：空单元格只能在位置0（开头）插入"
                return result
            new_text = insert_text
        else:
            # 检查插入位置有效性
            max_pos = len(original_text)
            if insert_pos > max_pos:
                result["error_message"] = f"错误：插入位置应在0至{max_pos}之间（当前值：{insert_pos}）"
                return result
            
            # 执行插入操作
            left_part = original_text[:insert_pos]
            right_part = original_text[insert_pos:]
            new_text = left_part + insert_text + right_part
        
        # 更新单元格并保存
        cell.value = new_text
        wb.save(file_save_path)
        
        result.update({"success": True, "result_text": new_text})
        
    except Exception as e:
        result["error_message"] = f"系统错误：{str(e)}"
    
    return result

from openpyxl import Workbook
def create_excel_function(input_path):
    wb = Workbook()
    # 保存为Excel文件
    wb.save(input_path)
    
def delete_single_cell_text_function(file_path,file_save_path, sheet_name, cell_address):
    """
    删除Excel单个单元格中指定区间的文字（索引从1开始）
    
    参数说明：
    - file_path:    Excel文件路径（不存在时报错）
    - sheet_name:    工作表名称
    - cell_address:  目标单元格地址（如"A1"，必须为单个单元格）
    - start_pos:     起始删除位置（包含）
    - end_pos:       结束删除位置（包含）
    
    返回值：
    - 字典包含三个键：
      - success:      操作是否成功（True/False）
      - error_message: 错误描述（成功时为空）
      - result_text:   删除后的文本（失败时为空）
    """
    result = {"success": False, "error_message": "", "result_text": ""}
    
    try:
        # 加载Excel文件（文件不存在时报错，参考网页3）
        wb = openpyxl.load_workbook(file_path)
        
        # 获取工作表
        if sheet_name not in wb.sheetnames:
            result["error_message"] = f"工作表 '{sheet_name}' 不存在"
            return result
        sheet = wb[sheet_name]
        
        # 检查单元格地址是否为单个单元格（参考网页6）
        if ':' in cell_address:
            result["error_message"] = "错误：必须指定单个单元格（如'A1'）"
            return result
        
        # 获取单元格并处理空值
        cell = sheet[cell_address]
        original_text = str(cell.value) if cell.value is not None else ""
        
        # 空单元格检查（参考网页6）
        if original_text == "":
            result["error_message"] = "错误：单元格为空，无法删除"
            return result
        
        # 参数有效性检查（索引从1开始）
        text_length = len(original_text)
        start_pos = 1
        end_pos = text_length
        # if start_pos < 1 or end_pos > text_length or start_pos > end_pos:
        #     result["error_message"] = f"错误：删除位置应在1至{text_length}之间，且起始位置≤结束位置"
        #     return result
        
        # 执行删除操作（参考网页3的字符串切片逻辑）
        new_text = original_text[:start_pos-1] + original_text[end_pos:]
        cell.value = new_text
        wb.save(file_save_path)
        
        result.update({"success": True, "result_text": new_text})
    
    except FileNotFoundError:
        result["error_message"] = f"错误：文件 '{file_path}' 不存在"
    except Exception as e:
        result["error_message"] = f"系统错误：{str(e)}"
    
    return result
