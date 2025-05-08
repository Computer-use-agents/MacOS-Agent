import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import os
from typing import List, Optional, Union
from matplotlib import rcParams

def plot_excel_data(
    file_path: str,
    sheet_name: str,
    chart_name: str,
    data_type: str,
    x_axis_label: str,
    y_axis_label: str,
    y_data_cols: List[Union[str, int]],
    # 可选参数
    output_sheet_name: Optional[str] = None,
    image_temp_path: str = "temp_chart.png",
    image_width: int = 800,
    image_height: int = 400,
    chart_offset_rows: int = 2,
    # 图表配置参数
    title: Optional[str] = None,
    x_data_col: Optional[Union[str, int]] = None,
    chart_type: str = 'line',
    has_header: bool = True,
    show_legend: bool = True,
    show_grid: bool = False
) -> dict[str, Union[bool, str]]:
    """
    读取Excel表格数据并绘制图表，然后将图表保存回Excel文件中
    
    参数:
        file_path (str): Excel文件路径
        sheet_name (str): 数据所在工作表名称
        chart_name (str): 图表名称
        data_type (str): 数据类型('1D'或'2D')
        x_axis_label (str): X轴标签
        y_axis_label (str): Y轴标签
        y_data_cols (List[Union[str, int]]): Y数据列名或索引列表
        
        output_sheet_name (str, optional): 输出工作表名称
        image_temp_path (str, optional): 临时图像保存路径
        image_width (int, optional): 图像宽度(像素)
        image_height (int, optional): 图像高度(像素)
        chart_offset_rows (int, optional): 图表与表格之间的空行数
        
        title (str, optional): 图表标题(默认使用chart_name)
        x_data_col (Union[str, int], optional): X数据列名或索引(2D数据必需)
        chart_type (str, optional): 图表类型('line', 'bar', 'scatter', 'pie')
        has_header (bool, optional): 数据是否有表头
        show_legend (bool, optional): 是否显示图例
        show_grid (bool, optional): 是否显示网格
    
    返回:
        Dict[str, Union[bool, str]]: 包含'success'和'message'的执行结果
    """
    result = {'success': False, 'message': ''}
    
    try:
        # 参数验证
        if data_type not in ['1D', '2D']:
            raise ValueError("data_type必须是'1D'或'2D'")
        
        if data_type == '2D' and x_data_col is None:
            raise ValueError("2D数据需要指定x_data_col参数")
        
        if not y_data_cols:
            raise ValueError("y_data_cols不能为空")
        
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Zen Hei']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        
        # 获取工作表
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        ws = wb[sheet_name]
        
        # 读取数据并确定数据范围
        data = list(ws.values)
        max_data_row = len(data)
        max_data_col = max(len(row) for row in data) if data else 0
        
        # 获取工作表实际使用的最大行和列
        if ws.max_row > max_data_row:
            max_data_row = ws.max_row
        if ws.max_column > max_data_col:
            max_data_col = ws.max_column
        
        # 创建图表
        dpi = 100
        fig_width = image_width / dpi
        fig_height = image_height / dpi
        plt.figure(figsize=(fig_width, fig_height), dpi=dpi)
        
        # 处理列索引/名称
        def get_col_index(col_spec: Union[str, int], header_row: Optional[tuple] = None) -> int:
            if isinstance(col_spec, int):
                return col_spec - 1  # 转换为0-based索引
            elif isinstance(col_spec, str) and header_row:
                try:
                    return header_row.index(col_spec)
                except ValueError:
                    raise ValueError(f"列名 '{col_spec}' 不存在于表头")
            else:
                raise ValueError("无效的列指定方式 - 需要列索引(整数)或列名(字符串且数据有表头)")
        
        header = data[0] if has_header else None
        start_row = 1 if has_header else 0
        
        # 根据数据类型处理数据
        if data_type == '1D':
            y_indices = [get_col_index(col, header) for col in y_data_cols]
            y_data = [[row[idx] for idx in y_indices] for row in data[start_row:]]
            x_data = list(range(1, len(y_data) + 1))
            
        else:  # '2D'
            x_index = get_col_index(x_data_col, header)
            y_indices = [get_col_index(col, header) for col in y_data_cols]
            x_data = [row[x_index] for row in data[start_row:]]
            y_data = [[row[idx] for idx in y_indices] for row in data[start_row:]]
        
        # 绘制图表
        if chart_type == 'line':
            for i, col_idx in enumerate(y_indices):
                label = header[col_idx] if has_header else f'Series {i+1}'
                plt.plot(x_data, [y[i] for y in y_data], label=label)
        elif chart_type == 'bar':
            width = 0.8 / len(y_indices)
            for i, col_idx in enumerate(y_indices):
                label = header[col_idx] if has_header else f'Series {i+1}'
                pos = [x + i * width for x in range(len(x_data))] if isinstance(x_data[0], str) else x_data
                plt.bar(pos, [y[i] for y in y_data], width=width, label=label)
        elif chart_type == 'scatter':
            for i, col_idx in enumerate(y_indices):
                label = header[col_idx] if has_header else f'Series {i+1}'
                plt.scatter(x_data, [y[i] for y in y_data], label=label)
        elif chart_type == 'pie':
            if len(y_indices) > 1:
                raise ValueError("饼图只能显示一个数据系列")
            plt.pie([y[0] for y in y_data], labels=x_data if isinstance(x_data[0], str) else None, autopct='%1.1f%%')
        else:
            raise ValueError(f"不支持的图表类型: {chart_type}")
        
        # 设置图表属性
        plt.xlabel(x_axis_label)
        plt.ylabel(y_axis_label)
        plt.title(title if title else chart_name)
        
        if show_legend and chart_type != 'pie':
            plt.legend()
        
        if show_grid:
            plt.grid(True)
        
        # 调整布局
        plt.tight_layout()
        
        # 保存临时图像
        plt.savefig(image_temp_path, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        # 将图像插入Excel
        output_ws = wb[output_sheet_name] if output_sheet_name and output_sheet_name in wb.sheetnames else ws
        
        img = Image(image_temp_path)
        chart_start_row = max_data_row + chart_offset_rows + 1
        output_ws.add_image(img, f'A{chart_start_row}')
        
        # 自动调整列宽
        if max_data_col < 10:
            for col in range(1, max_data_col + 1):
                output_ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存工作簿
        wb.save(file_path)
        
        # 删除临时图像
        if os.path.exists(image_temp_path):
            os.remove(image_temp_path)
        
        result['success'] = True
        result['message'] = f"图表 '{chart_name}' 已成功创建并保存到工作表 '{output_sheet_name or sheet_name}'"
    
    except Exception as e:
        result['message'] = f"错误: {str(e)}"
        if 'image_temp_path' in locals() and os.path.exists(image_temp_path):
            os.remove(image_temp_path)
    
    return result

result = plot_excel_data(
    file_path="test_new.xlsx",
    sheet_name="Sheet1",
    chart_name="产品对比",
    data_type="1D",
    x_axis_label="产品",
    y_axis_label="销量",
    y_data_cols=["Q1", "Q2", "Q3"],
    x_data_col="产品名称",
    chart_type="line",
    output_sheet_name="产品图表",
    chart_offset_rows=3
)
# print(result)